from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import json
import os
import xlwt
from openpyxl import load_workbook
import xlwt
import xlrd
from openpyxl import load_workbook
import zipfile
import tempfile
import os
import traceback

app = Flask(__name__)

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/indexconversor")
def indexconversor():
    return render_template("indexconversor.html")

@app.route("/converter-xls", methods=["GET", "POST"])
def converter_xls():
    """Rota para converter arquivo XLSX para XLS preservando fórmulas"""
    if request.method == "GET":
        return render_template("converter.html")
    
    if request.method == "POST":
        try:
            if 'arquivo' not in request.files:
                return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
            
            arquivo = request.files['arquivo']
            if arquivo.filename == '':
                return jsonify({'erro': 'Nenhum arquivo foi selecionado'}), 400
            
            # Salva o arquivo temporariamente para análise
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                arquivo.save(temp_file.name)
                temp_path = temp_file.name
            
            try:
                # Primeira verificação: é um arquivo ZIP válido?
                if not zipfile.is_zipfile(temp_path):
                    # Pode ser um arquivo XLS disfarçado de XLSX
                    try:
                        # Tenta abrir como XLS
                        xlrd_book = xlrd.open_workbook(temp_path)
                        return jsonify({'erro': 'Este arquivo parece ser XLS, não XLSX. Ele já está no formato desejado!'}), 400
                    except:
                        return jsonify({'erro': 'Arquivo não é um Excel válido (nem XLSX nem XLS)'}), 400
                
                # Segunda verificação: consegue abrir com openpyxl?
                # IMPORTANTE: data_only=False para preservar fórmulas
                try:
                    wb_xlsx = load_workbook(temp_path, read_only=True, data_only=False)
                    ws_xlsx = wb_xlsx.active
                except Exception as e:
                    return jsonify({'erro': f'Erro ao abrir arquivo XLSX: {str(e)}'}), 400
                
                # Cria um novo workbook XLS usando xlwt
                wb_xls = xlwt.Workbook()
                ws_xls = wb_xls.add_sheet("Planilha", cell_overwrite_ok=True)
                
                # Define estilos básicos para xlwt
                header_style = xlwt.XFStyle()
                header_font = xlwt.Font()
                header_font.bold = True
                header_style.font = header_font
                
                # Estilo para números
                number_style = xlwt.XFStyle()
                number_style.num_format_str = '0.00'
                
                # Contador de linhas e colunas processadas
                rows_processed = 0
                cols_processed = 0
                formulas_converted = 0
                
                # Copia os dados do XLSX para XLS PRESERVANDO FÓRMULAS
                # MUDANÇA PRINCIPAL: usar iter_rows() sem values_only para acessar objetos Cell
                for row_idx, row in enumerate(ws_xlsx.iter_rows()):
                    if row_idx >= 65535:  # Limite do XLS
                        break
                    
                    row_has_data = False
                    for col_idx, cell in enumerate(row):
                        if col_idx >= 255:  # Limite do XLS
                            break
                        
                        if cell.value is not None:
                            row_has_data = True
                            cols_processed = max(cols_processed, col_idx + 1)
                            
                            try:
                                # NOVA LÓGICA: Verifica se a célula contém fórmula
                                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                    # É uma fórmula - preserva no XLS
                                    formula_xlsx = str(cell.value)
                                    
                                    # Remove o '=' inicial se presente (xlwt adiciona automaticamente)
                                    if formula_xlsx.startswith('='):
                                        formula_xls = formula_xlsx[1:]
                                    else:
                                        formula_xls = formula_xlsx
                                    
                                    # Escreve a fórmula no XLS usando xlwt.Formula
                                    ws_xls.write(row_idx, col_idx, xlwt.Formula(formula_xls))
                                    formulas_converted += 1
                                    
                                else:
                                    # Não é fórmula - trata como valor normal
                                    if row_idx == 0:
                                        # Aplica estilo de cabeçalho na primeira linha
                                        ws_xls.write(row_idx, col_idx, str(cell.value), header_style)
                                    else:
                                        # Trata diferentes tipos de dados
                                        if isinstance(cell.value, (int, float)):
                                            ws_xls.write(row_idx, col_idx, cell.value, number_style)
                                        elif isinstance(cell.value, bool):
                                            ws_xls.write(row_idx, col_idx, str(cell.value))
                                        else:
                                            # Converte para string e limita o tamanho
                                            str_value = str(cell.value)[:32767]  # Limite do XLS
                                            ws_xls.write(row_idx, col_idx, str_value)
                                            
                            except Exception as cell_error:
                                # Se der erro em uma célula específica, escreve como string
                                try:
                                    ws_xls.write(row_idx, col_idx, str(cell.value)[:32767])
                                except:
                                    ws_xls.write(row_idx, col_idx, "ERRO_CONVERSAO")
                    
                    if row_has_data:
                        rows_processed = row_idx + 1
                
                # Fecha o workbook XLSX
                wb_xlsx.close()
                
                # Salva o arquivo XLS em memória
                output = BytesIO()
                wb_xls.save(output)
                output.seek(0)
                
                # Gera nome do arquivo baseado no original
                nome_original = arquivo.filename.rsplit('.', 1)[0]
                filename = f"{nome_original}_convertido.xls"
                
                # Log de informações da conversão
                print(f"Conversão XLSX->XLS concluída: {rows_processed} linhas, {cols_processed} colunas, {formulas_converted} fórmulas preservadas")
                
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.ms-excel'
                )
                
            finally:
                # Remove o arquivo temporário
                try:
                    os.unlink(temp_path)
                except:
                    pass
            
        except Exception as e:
            return jsonify({'erro': f'Erro inesperado: {str(e)}'}), 500

@app.route("/converter-xlsx", methods=["GET", "POST"])
def converter_xlsx():
    """Rota para converter arquivo XLS para XLSX"""
    if request.method == "GET":
        return render_template("converter_xlsx.html")
    
    if request.method == "POST":
        try:
            if 'arquivo' not in request.files:
                return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
            
            arquivo = request.files['arquivo']
            if arquivo.filename == '':
                return jsonify({'erro': 'Nenhum arquivo foi selecionado'}), 400
            
            # Verifica se é um arquivo XLS
            if not arquivo.filename.lower().endswith('.xls'):
                return jsonify({'erro': 'Arquivo deve ser uma planilha XLS (.xls)'}), 400
            
            # Salva o arquivo temporariamente
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_file:
                arquivo.save(temp_file.name)
                temp_path = temp_file.name
            
            try:
                # Tenta abrir o arquivo XLS
                try:
                    wb_xls = xlrd.open_workbook(temp_path, formatting_info=True)
                    ws_xls = wb_xls.sheet_by_index(0)
                except Exception as e:
                    return jsonify({'erro': f'Erro ao abrir arquivo XLS: {str(e)}'}), 400
                
                # Cria um novo workbook XLSX
                wb_xlsx = Workbook()
                ws_xlsx = wb_xlsx.active
                ws_xlsx.title = "Convertido"
                
                # Estilos para XLSX
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0056A4", end_color="0056A4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Contador de dados processados
                rows_processed = 0
                cols_processed = 0
                
                # Tipos de célula no xlrd
                XL_CELL_EMPTY = 0
                XL_CELL_TEXT = 1
                XL_CELL_NUMBER = 2
                XL_CELL_DATE = 3
                XL_CELL_BOOLEAN = 4
                XL_CELL_ERROR = 5
                
                # Copia os dados do XLS para XLSX
                for row_idx in range(ws_xls.nrows):
                    row_has_data = False
                    for col_idx in range(ws_xls.ncols):
                        cell_xls = ws_xls.cell(row_idx, col_idx)
                        cell_xlsx = ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1)
                        
                        try:
                            if cell_xls.ctype == XL_CELL_EMPTY:
                                continue
                            elif cell_xls.ctype == XL_CELL_TEXT:
                                cell_xlsx.value = cell_xls.value
                                row_has_data = True
                            elif cell_xls.ctype == XL_CELL_NUMBER:
                                cell_xlsx.value = cell_xls.value
                                row_has_data = True
                            elif cell_xls.ctype == XL_CELL_DATE:
                                cell_xlsx.value = cell_xls.value
                                row_has_data = True
                            elif cell_xls.ctype == XL_CELL_BOOLEAN:
                                cell_xlsx.value = bool(cell_xls.value)
                                row_has_data = True
                            elif cell_xls.ctype == XL_CELL_ERROR:
                                cell_xlsx.value = f"#ERROR#{cell_xls.value}"
                                row_has_data = True
                            else:
                                # Tipo desconhecido ou fórmula (xlrd 2.0+ não suporta fórmulas)
                                cell_xlsx.value = cell_xls.value
                                row_has_data = True
                            
                            # Aplica formatação para cabeçalho (primeira linha)
                            if row_idx == 0 and cell_xlsx.value:
                                cell_xlsx.font = header_font
                                cell_xlsx.fill = header_fill
                                cell_xlsx.alignment = header_alignment
                            
                            # Aplica borda
                            cell_xlsx.border = border
                            
                            if row_has_data:
                                cols_processed = max(cols_processed, col_idx + 1)
                            
                        except Exception as cell_error:
                            # Em caso de erro, escreve como string
                            cell_xlsx.value = str(cell_xls.value) if cell_xls.value else ""
                            cell_xlsx.border = border
                    
                    if row_has_data:
                        rows_processed = row_idx + 1
                
                # Ajusta largura das colunas
                for col in range(1, cols_processed + 1):
                    column_letter = get_column_letter(col)
                    ws_xlsx.column_dimensions[column_letter].width = 15
                
                # Salva o arquivo XLSX em memória
                output = BytesIO()
                wb_xlsx.save(output)
                output.seek(0)
                
                # Gera nome do arquivo baseado no original
                nome_original = arquivo.filename.rsplit('.', 1)[0]
                filename = f"{nome_original}_convertido.xlsx"
                
                # Log de informações da conversão
                print(f"Conversão XLS->XLSX concluída: {rows_processed} linhas, {cols_processed} colunas")
                
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            finally:
                # Remove o arquivo temporário
                try:
                    os.unlink(temp_path)
                except:
                    pass
            
        except Exception as e:
            return jsonify({'erro': f'Erro inesperado: {str(e)}'}), 500


    """Versão alternativa usando pandas para conversão (SEM preservação de fórmulas)"""
    try:
        import pandas as pd
        
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            return jsonify({'erro': 'Nenhum arquivo foi selecionado'}), 400
        
        # Salva temporariamente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            arquivo.save(temp_file.name)
            temp_path = temp_file.name
        
        try:
            # Lê o arquivo XLSX com pandas (NOTA: pandas não preserva fórmulas)
            df = pd.read_excel(temp_path, engine='openpyxl')
            
            # Cria arquivo XLS temporário
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_xls:
                temp_xls_path = temp_xls.name
            
            # Salva como XLS
            df.to_excel(temp_xls_path, engine='xlwt', index=False)
            
            # Lê o arquivo XLS gerado
            with open(temp_xls_path, 'rb') as f:
                output = BytesIO(f.read())
            
            # Remove arquivos temporários
            os.unlink(temp_path)
            os.unlink(temp_xls_path)
            
            # Gera nome do arquivo
            nome_original = arquivo.filename.rsplit('.', 1)[0]
            filename = f"{nome_original}_convertido_pandas.xls"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.ms-excel'
            )
            
        except Exception as e:
            # Limpa arquivos temporários em caso de erro
            try:
                os.unlink(temp_path)
            except:
                pass
            return jsonify({'erro': f'Erro na conversão com pandas: {str(e)}'}), 500
            
    except ImportError:
        return jsonify({'erro': 'Pandas não está instalado. Use a rota principal que preserva fórmulas.'}), 500
    except Exception as e:
        return jsonify({'erro': f'Erro inesperado: {str(e)}'}), 500

@app.route("/importar", methods=["POST"])
def importar():
    """Importa dados de uma planilha existente para edição - VERSÃO CORRIGIDA"""
    try:
        print("=== INÍCIO DA IMPORTAÇÃO ===")
        
        if 'arquivo' not in request.files:
            print("Erro: Nenhum arquivo foi enviado")
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            print("Erro: Nenhum arquivo foi selecionado")
            return jsonify({'erro': 'Nenhum arquivo foi selecionado'}), 400
        
        print(f"Arquivo recebido: {arquivo.filename}")
        
        # Verifica se é um arquivo Excel
        if not arquivo.filename.lower().endswith(('.xlsx', '.xls')):
            print("Erro: Arquivo não é Excel")
            return jsonify({'erro': 'Arquivo deve ser uma planilha Excel (.xlsx ou .xls)'}), 400
        
        # Carrega a planilha
        try:
            print("Tentando carregar planilha...")
            wb = load_workbook(arquivo)
            ws = wb.active
            print(f"Planilha carregada: {ws.title}, {ws.max_row} linhas, {ws.max_column} colunas")
        except Exception as e:
            print(f"Erro ao carregar planilha: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return jsonify({'erro': f'Erro ao carregar planilha: {str(e)}'}), 400
        
        dados_importados = []
        linhas_processadas = 0
        linhas_com_dados = 0
        erros_por_linha = []
        
        # Examinar a estrutura primeiro
        print("=== EXAMINANDO ESTRUTURA ===")
        print("Cabeçalhos (linha 1):")
        for col in range(1, min(10, ws.max_column + 1)):
            valor = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {repr(valor)}")
        
        # Lê os dados a partir da linha 3 (pula cabeçalho e totais)
        print("=== PROCESSANDO DADOS ===")
        for linha in range(3, ws.max_row + 1):
            linhas_processadas += 1
            print(f"Processando linha {linha}")
            
            try:
                # Verifica se a linha tem dados (pelo menos o código do produto)
                codigo = ws.cell(row=linha, column=2).value
                print(f"  Código (col 2): {repr(codigo)}")
                
                if not codigo or str(codigo).strip() == "":
                    print(f"  Linha {linha} sem código válido, pulando...")
                    continue
                
                linhas_com_dados += 1
                print(f"  Linha {linha} tem dados, processando...")
                
                # Função auxiliar para obter valor seguro
                def get_safe_value(row, col, tipo='str', default=None):
                    try:
                        valor = ws.cell(row=row, column=col).value
                        if valor is None:
                            return default if default is not None else ('' if tipo == 'str' else 0)
                        
                        if tipo == 'str':
                            return str(valor).strip()
                        elif tipo == 'float':
                            return float(valor) if valor != '' else 0
                        elif tipo == 'int':
                            return int(valor) if valor != '' else 0
                        else:
                            return valor
                    except Exception as e:
                        print(f"    Erro ao obter valor linha {row}, col {col}: {e}")
                        return default if default is not None else ('' if tipo == 'str' else 0)
                
                # Criar item com tratamento de erro individual para cada campo
                item = {}
                
                # Campos obrigatórios
                item['codigo'] = get_safe_value(linha, 2, 'str', '')
                item['cfop'] = get_safe_value(linha, 3, 'str', '')
                item['quantidade'] = get_safe_value(linha, 4, 'float', 0)
                item['valorUnitario'] = get_safe_value(linha, 5, 'float', 0)
                
                # Campos opcionais
                item['seguro'] = get_safe_value(linha, 7, 'float', 0)
                item['frete'] = get_safe_value(linha, 8, 'float', 0)
                item['desconto'] = get_safe_value(linha, 9, 'float', 0)
                item['outrasDespesas'] = get_safe_value(linha, 10, 'float', 0)
                item['numeroDI'] = get_safe_value(linha, 11, 'str', '')
                item['dataRegistro'] = get_safe_value(linha, 12, 'str', '')
                item['codigoExportador'] = get_safe_value(linha, 13, 'str', '')
                item['viaTransporte'] = get_safe_value(linha, 14, 'int', 0)
                item['valorAFRMM'] = get_safe_value(linha, 15, 'float', 0)
                item['desembaracoUF'] = get_safe_value(linha, 16, 'str', '')
                item['desembaracoLocal'] = get_safe_value(linha, 17, 'str', '')
                item['desembaracoData'] = get_safe_value(linha, 18, 'str', '')
                item['adicao'] = get_safe_value(linha, 19, 'int', 0)
                item['itemAdicao'] = get_safe_value(linha, 20, 'int', 0)
                item['codigoFabricante'] = get_safe_value(linha, 21, 'str', '')
                item['percentualII'] = get_safe_value(linha, 22, 'float', 0)
                item['baseII'] = get_safe_value(linha, 23, 'float', 0)
                item['despesasAduaneiras'] = get_safe_value(linha, 25, 'float', 0)
                item['valorIOF'] = get_safe_value(linha, 26, 'float', 0)
                item['cstIPI'] = get_safe_value(linha, 27, 'str', '00')
                item['percentualIPI'] = get_safe_value(linha, 28, 'float', 0)
                item['baseIPI'] = get_safe_value(linha, 29, 'float', 0)
                item['cstPIS'] = get_safe_value(linha, 31, 'int', 0)
                item['percentualPIS'] = get_safe_value(linha, 32, 'float', 0)
                item['basePIS'] = get_safe_value(linha, 33, 'float', 0)
                item['cstCOFINS'] = get_safe_value(linha, 35, 'int', 0)
                item['percentualCOFINS'] = get_safe_value(linha, 36, 'float', 0)
                item['baseCOFINS'] = get_safe_value(linha, 37, 'float', 0)
                item['cstICMS'] = get_safe_value(linha, 39, 'int', 0)
                item['percentualICMS'] = get_safe_value(linha, 40, 'float', 0)
                item['percentualRedICMS'] = get_safe_value(linha, 41, 'float', 0)
                item['baseICMS'] = get_safe_value(linha, 42, 'float', 0)
                item['valorICMSST'] = get_safe_value(linha, 44, 'float', 0)
                
                dados_importados.append(item)
                print(f"  Item criado com sucesso: código={item['codigo']}")
                
            except Exception as e:
                erro_msg = f"Erro na linha {linha}: {str(e)}"
                print(f"  ERRO: {erro_msg}")
                erros_por_linha.append(erro_msg)
                # Continua processando outras linhas mesmo com erro
        
        print("=== RESULTADO DA IMPORTAÇÃO ===")
        print(f"Linhas processadas: {linhas_processadas}")
        print(f"Linhas com dados: {linhas_com_dados}")
        print(f"Itens importados: {len(dados_importados)}")
        print(f"Erros encontrados: {len(erros_por_linha)}")
        
        # Verificar se encontrou dados válidos
        if len(dados_importados) == 0:
            if linhas_com_dados == 0:
                mensagem_erro = "A planilha não contém dados válidos. Verifique se:\n"
                mensagem_erro += "- Os dados estão nas linhas corretas (a partir da linha 3)\n"
                mensagem_erro += "- A coluna 2 contém os códigos dos produtos\n"
                mensagem_erro += "- O arquivo não está vazio ou corrompido"
                print(f"Erro: {mensagem_erro}")
                return jsonify({'erro': mensagem_erro}), 400
            else:
                mensagem_erro = f"Encontradas {linhas_com_dados} linhas com códigos, mas nenhuma pôde ser processada."
                if erros_por_linha:
                    mensagem_erro += f"\nErros encontrados:\n" + "\n".join(erros_por_linha[:5])
                print(f"Erro: {mensagem_erro}")
                return jsonify({'erro': mensagem_erro}), 400
        
        # Sucesso
        resultado = {
            'dados': dados_importados, 
            'sucesso': True,
            'estatisticas': {
                'linhas_processadas': linhas_processadas,
                'linhas_com_dados': linhas_com_dados,
                'itens_importados': len(dados_importados),
                'erros': len(erros_por_linha)
            }
        }
        
        if erros_por_linha:
            resultado['avisos'] = erros_por_linha[:10]  # Máximo 10 avisos
        
        print("Importação concluída com sucesso!")
        return jsonify(resultado)
        
    except Exception as e:
        erro_msg = f'Erro inesperado ao importar planilha: {str(e)}'
        print(f"ERRO CRÍTICO: {erro_msg}")
        print(f"Traceback completo: {traceback.format_exc()}")
        return jsonify({'erro': erro_msg}), 500

@app.route("/exportar", methods=["POST"])
def exportar():
    dados = request.get_json()
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # Define os cabeçalhos conforme especificação
    cabecalhos = [
        "item", "codigo", "CFOP", "quantidade", "valor unitário", "valor total",
        "seguro", "frete", "desconto", "outras despesas", "Número DI/DSI/DA",
        "Data registro", "Código do Exportador", "Via de transporte", "Valor AFRMM",
        "Desembaraço (UF)", "Desembaraço (Local)", "Desembaraço (Data)", "adição",
        "item adição", "Código Fabricante", "%II", "Base II", "Valor II",
        "Despesas aduaneiras", "Valor IOF", "CST IPI", "%IPI", "Base IPI",
        "Valor IPI", "CST PIS", "%PIS", "Base PIS", "Valor PIS", "CST COFINS",
        "%COFINS", "Base COFINS", "Valor COFINS", "CST ICMS", "%ICMS", "%Red ICMS",
        "Base ICMS", "Valor ICMS", "Valor ICMS ST"
    ]

    # Linha 1: Cabeçalhos
    for col, cabecalho in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=cabecalho)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0056A4", end_color="0056A4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Linha 2: Totais (fórmulas de soma para colunas específicas)
    colunas_soma = {
        6: "F",   # valor total
        7: "G",   # seguro
        8: "H",   # frete
        9: "I",   # desconto
        10: "J",  # outras despesas
        15: "O",  # Valor AFRMM
        23: "W",  # Base II
        24: "X",  # Valor II
        25: "Y",  # Despesas aduaneiras
        26: "Z",  # Valor IOF
        29: "AC", # Base IPI
        30: "AD", # Valor IPI
        33: "AG", # Base PIS
        34: "AH", # Valor PIS
        37: "AK", # Base COFINS
        38: "AL", # Valor COFINS
        42: "AP", # Base ICMS
        43: "AQ", # Valor ICMS
        44: "AR"  # Valor ICMS ST
    }

    for col in range(1, len(cabecalhos) + 1):
        if col in colunas_soma:
            letra_col = colunas_soma[col]
            formula = f"=SUM({letra_col}3:{letra_col}16)"
            cell = ws.cell(row=2, column=col, value=formula)
        else:
            cell = ws.cell(row=2, column=col, value="")
        
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # A partir da linha 3: Dados dos itens
    if dados:
        for linha_idx, item in enumerate(dados, 3):
            # Mapear os dados do formulário para as colunas corretas
            dados_linha = [
                linha_idx - 2,  # item (numeração sequencial)
                item.get('codigo', ''),
                item.get('cfop', ''),
                float(item.get('quantidade', 0)) if item.get('quantidade') else 0,
                float(item.get('valorUnitario', 0)) if item.get('valorUnitario') else 0,
                f"=D{linha_idx}*E{linha_idx}",  # valor total (fórmula)
                float(item.get('seguro', 0)) if item.get('seguro') else 0,
                float(item.get('frete', 0)) if item.get('frete') else 0,
                float(item.get('desconto', 0)) if item.get('desconto') else 0,
                float(item.get('outrasDespesas', 0)) if item.get('outrasDespesas') else 0,
                item.get('numeroDI', ''),
                item.get('dataRegistro', ''),
                item.get('codigoExportador', ''),
                int(item.get('viaTransporte', 0)) if item.get('viaTransporte') else 0,
                float(item.get('valorAFRMM', 0)) if item.get('valorAFRMM') else 0,
                item.get('desembaracoUF', ''),
                item.get('desembaracoLocal', ''),
                item.get('desembaracoData', ''),
                int(item.get('adicao', 0)) if item.get('adicao') else 0,
                int(item.get('itemAdicao', 0)) if item.get('itemAdicao') else 0,
                item.get('codigoFabricante', ''),
                float(item.get('percentualII', 0)) if item.get('percentualII') else 0,
                float(item.get('baseII', 0)) if item.get('baseII') else 0,  # Base II agora é valor numérico
                f"=ROUND(W{linha_idx}*V{linha_idx}/100,2)",  # Valor II (fórmula corrigida)
                float(item.get('despesasAduaneiras', 0)) if item.get('despesasAduaneiras') else 0,
                float(item.get('valorIOF', 0)) if item.get('valorIOF') else 0,
                int(item.get('cstIPI', 0)) if item.get('cstIPI') else 0,
                float(item.get('percentualIPI', 0)) if item.get('percentualIPI') else 0,
                float(item.get('baseIPI', 0)) if item.get('baseIPI') else 0,
                f"=ROUND(AC{linha_idx}*AB{linha_idx}/100,2)",  # Valor IPI (fórmula)
                int(item.get('cstPIS', 0)) if item.get('cstPIS') else 0,
                float(item.get('percentualPIS', 0)) if item.get('percentualPIS') else 0,
                float(item.get('basePIS', 0)) if item.get('basePIS') else 0,
                f"=ROUND(AG{linha_idx}*AF{linha_idx}/100,2)",  # Valor PIS (fórmula)
                int(item.get('cstCOFINS', 0)) if item.get('cstCOFINS') else 0,
                float(item.get('percentualCOFINS', 0)) if item.get('percentualCOFINS') else 0,
                float(item.get('baseCOFINS', 0)) if item.get('baseCOFINS') else 0,
                f"=ROUND(AK{linha_idx}*AJ{linha_idx}/100,2)",  # Valor COFINS (fórmula)
                int(item.get('cstICMS', 0)) if item.get('cstICMS') else 0,
                float(item.get('percentualICMS', 0)) if item.get('percentualICMS') else 0,
                float(item.get('percentualRedICMS', 0)) if item.get('percentualRedICMS') else 0,
                float(item.get('baseICMS', 0)) if item.get('baseICMS') else 0,
                f"=ROUND(AP{linha_idx}*AN{linha_idx}/100,2)",  # Valor ICMS (fórmula)
                float(item.get('valorICMSST', 0)) if item.get('valorICMSST') else 0
            ]

            for col, valor in enumerate(dados_linha, 1):
                cell = ws.cell(row=linha_idx, column=col, value=valor)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Formatação especial para números
                if isinstance(valor, (int, float)) and valor != 0:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(valor, str) and valor.startswith('='):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Ajustar largura das colunas
    for col in range(1, len(cabecalhos) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Salva em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Retorna o arquivo para download
    filename = f"planilha_sisloc_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
    return send_file(
        output, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)

